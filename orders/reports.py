from io import BytesIO
import os
from django.conf import settings
from django.db.models import Sum, Avg, Min, Max, Count
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from .models import Order, OrderItem

REPORT_DIR = os.path.join(settings.MEDIA_ROOT, "reports")
os.makedirs(REPORT_DIR, exist_ok=True)


def apply_excel_style(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if row_idx == 1 else "left")

            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_length + 4, 14)

    ws.freeze_panes = "A2"


def set_money_format(cell):
    """Применяет денежный формат к ячейке (русская локаль)."""
    cell.number_format = '#,##0.00'


def generate_report_content(report_type, date_from, date_to, file_format="xlsx"):
    """
    Генератор отчётов Excel: продажи, средний чек, топ товаров.
    """
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active

    # ===================== ПРОДАЖИ ПО МЕСЯЦАМ =====================
    if report_type == "sales_by_month":
        ws.title = "Продажи и заказы"

        ws.append([
            "Год",
            "Месяц",
            "Количество заказов",
            "Сумма продаж",
            "Средний чек",
            "Максимальный заказ",
            "Минимальный заказ"
        ])

        stats = (
            Order.objects.filter(created_at__date__range=[date_from, date_to])
            .values("created_at__year", "created_at__month")
            .annotate(
                total=Sum("total_price"),
                avg=Avg("total_price"),
                min=Min("total_price"),
                max=Max("total_price"),
                count=Count("id")
            )
            .order_by("created_at__year", "created_at__month")
        )

        for row in stats:
            ws.append([
                row["created_at__year"],
                row["created_at__month"],
                row["count"],
                row["total"] or 0,
                row["avg"] or 0,
                row["max"] or 0,
                row["min"] or 0,
            ])

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=7):
            for cell in row:
                set_money_format(cell)

        if ws.max_row > 1:
            ws.append([])
            total_row = ws.max_row + 1
            ws.cell(row=total_row, column=3, value="Итого:")
            ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row - 2})")
            set_money_format(ws.cell(row=total_row, column=4))

        ws.append([])
        ws.append(["СПИСОК ЗАКАЗОВ"])
        ws.append([])

        ws.append([
            "ID",
            "Номер заказа",
            "Дата",
            "Пользователь",
            "Статус",
            "Страна",
            "Доставка",
            "Сумма заказа",
            "Стоимость доставки",
            "Итог"
        ])

        orders = Order.objects.filter(created_at__date__range=[date_from, date_to]).select_related("user")

        if orders.exists():
            start_row = ws.max_row + 1
            for order in orders:
                ws.append([
                    order.id,
                    order.order_number,
                    order.created_at.strftime("%Y-%m-%d"),
                    str(order.user),
                    order.get_status_display(),
                    order.get_country_display(),
                    order.get_delivery_method_display(),
                    float(order.total_price),
                    float(order.delivery_price),
                    float(order.total_price + order.delivery_price),
                ])

            for row in ws.iter_rows(min_row=start_row, min_col=8, max_col=10):
                for cell in row:
                    set_money_format(cell)

    # ===================== СРЕДНИЙ ЧЕК =====================
    elif report_type == "average_check":
        ws.title = "Средний чек"
        orders = Order.objects.filter(created_at__date__range=[date_from, date_to])
        agg = orders.aggregate(
            avg=Avg("total_price"),
            min=Min("total_price"),
            max=Max("total_price")
        )
        total_orders = orders.count()

        ws.append(["Показатель", "Значение"])
        numeric_rows = {
            "Средний чек": agg["avg"],
            "Минимальный заказ": agg["min"],
            "Максимальный заказ": agg["max"],
        }

        for key, value in numeric_rows.items():
            ws.append([key, round(float(value or 0), 2)])

        ws.append(["Всего заказов", total_orders])

        max_item = (
            OrderItem.objects.filter(order__in=orders)
            .order_by("-price_snapshot")
            .first()
        )
        min_item = (
            OrderItem.objects.filter(order__in=orders)
            .order_by("price_snapshot")
            .first()
        )

        if max_item:
            ws.append(["Самая дорогая позиция", max_item.product_name])
            ws.append(["Цена", round(float(max_item.price_snapshot), 2)])
            ws.append(["Количество", max_item.quantity])

        if min_item:
            ws.append(["Самая дешёвая позиция", min_item.product_name])
            ws.append(["Цена", round(float(min_item.price_snapshot), 2)])
            ws.append(["Количество", min_item.quantity])

        for row in ws.iter_rows(min_row=2, max_col=2):
            val = row[1].value
            try:
                float(val)
                set_money_format(row[1])
            except (TypeError, ValueError):
                continue

    # ===================== ТОП ТОВАРОВ =====================
    elif report_type == "top_products":
        ws.title = "Популярные товары"
        ws.append(["ID товара", "Товар", "Цвет", "Размер", "Количество"])

        data = (
            OrderItem.objects.filter(order__created_at__date__range=[date_from, date_to])
            .values("variant__product__id", "product_name", "color", "size")
            .annotate(total_qty=Sum("quantity"))
            .order_by("-total_qty")
        )

        for row in data:
            ws.append([
                row["variant__product__id"],
                row["product_name"],
                row["color"],
                row["size"],
                row["total_qty"]
            ])

    apply_excel_style(ws)

    wb.save(buffer)
    buffer.seek(0)

    filename = f"report_{report_type}_{date_from}_to_{date_to}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)

    with open(filepath, "wb") as f:
        f.write(buffer.read())

    return os.path.join("reports", filename)