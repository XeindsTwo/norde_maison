from django.contrib import admin
from django.http import HttpResponse
from django.db.models import Count
from django.utils import timezone
from datetime import timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

from .models import Favorite


@admin.register(Favorite)
class FavoriteAdmin(admin.ModelAdmin):
    change_list_template = "admin/favorites_dashboard.html"

    list_display = ("product_verbose", "created_at")
    list_display_links = None

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            "product"
        ).filter(product__is_visible=True)

    def product_verbose(self, obj):
        return obj.product.name

    product_verbose.short_description = "Товар"

    def export_excel(self, queryset, period_text):

        stats = (
            queryset.values("product__name")
            .annotate(total=Count("id"))
            .order_by("-total")[:15]
        )

        wb = openpyxl.Workbook()
        ws = wb.active

        now = timezone.now()

        month_names = [
            "",
            "января", "февраля", "марта",
            "апреля", "мая", "июня",
            "июля", "августа", "сентября",
            "октября", "ноября", "декабря"
        ]

        date_str = f"{now.day} {month_names[now.month]} {now.year} г."

        ws.merge_cells("A1:B1")
        ws["A1"] = "Популярные избранные товары"
        ws["A2"] = f"Период: {period_text}"
        ws["A3"] = f"Дата отчёта: {date_str}"

        ws["A1"].font = Font(size=18, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A2"].font = Font(size=11)
        ws["A3"].font = Font(size=11)

        ws.column_dimensions["A"].width = 65
        ws.column_dimensions["B"].width = 25

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        fill_header = PatternFill(
            start_color="E9EDF4",
            end_color="E9EDF4",
            fill_type="solid"
        )

        header_row = 5
        headers = ["Товар", "Количество"]

        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)

            cell.font = Font(bold=True, size=12)
            cell.border = border
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        total_all = 0
        row = header_row + 1

        for item in stats:
            name = item["product__name"]
            total = item["total"]

            total_all += total

            name_cell = ws.cell(row=row, column=1, value=name)
            name_cell.alignment = Alignment(wrap_text=True, vertical="top")
            name_cell.border = border

            count_cell = ws.cell(row=row, column=2, value=total)
            count_cell.alignment = Alignment(horizontal="center")
            count_cell.border = border

            row += 1

        footer_row = row

        total_label = ws.cell(
            row=footer_row,
            column=1,
            value="Всего товаров"
        )
        total_label.font = Font(bold=True)

        total_value = ws.cell(
            row=footer_row,
            column=2,
            value=total_all
        )

        total_value.font = Font(bold=True)
        total_value.alignment = Alignment(horizontal="center")

        for col in (1, 2):
            cell = ws.cell(row=footer_row, column=col)
            cell.border = border
            cell.fill = PatternFill(
                start_color="F7F9FB",
                end_color="F7F9FB",
                fill_type="solid"
            )

        chart = PieChart()

        data = Reference(
            ws,
            min_col=2,
            min_row=header_row + 1,
            max_row=row - 1
        )

        labels = Reference(
            ws,
            min_col=1,
            min_row=header_row + 1,
            max_row=row - 1
        )

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True

        chart.legend = None

        chart.width = 20
        chart.height = 12

        ws.add_chart(chart, "D5")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = 'attachment; filename="popular_favorites.xlsx"'

        wb.save(response)

        return response

    def changelist_view(self, request, extra_context=None):

        if request.GET.get("export") == "1":

            period = request.GET.get("period", "all")

            queryset = self.get_queryset(request)

            if period == "week":
                queryset = queryset.filter(
                    created_at__gte=timezone.now() - timedelta(days=7)
                )

            elif period == "month":
                queryset = queryset.filter(
                    created_at__gte=timezone.now() - timedelta(days=30)
                )

            period_map = {
                "all": "весь период",
                "week": "неделя",
                "month": "месяц"
            }

            return self.export_excel(
                queryset,
                period_map.get(period, "весь период")
            )

        extra_context = extra_context or {}
        extra_context["info_text"] = "Выберите период отчёта и нажмите кнопку Excel"

        return super().changelist_view(request, extra_context=extra_context)
